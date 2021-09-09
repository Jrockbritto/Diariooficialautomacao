from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import time
import xlsxwriter
from openpyxl import load_workbook
import logging
from dateutil.parser import parse
from datetime import date
from datetime import datetime
import os


def get_all_links(driver):
    links = []
    titulo = []
    data = []
    table = driver.find_element_by_id('dgDocumentos')
    elements = table.find_elements_by_tag_name('tr')
    for i in range(1, len(elements)-1):  # eliminando cabeçalho e footer da table
        href = elements[i].find_element_by_tag_name('a').get_attribute("href")
        d = href.split("data=", 1)[1]
        n = d.split("&", 1)[0]
        l = href.split("doc=", 1)[1]
        links.append(
            'http://diariooficial.rn.gov.br/dei/dorn3/documentos/00000001/'+n+'/'+l+'.htm')
        titulo.append(elements[i].find_element_by_xpath(
            './/td[3]').get_attribute('innerHTML'))
        data.append(elements[i].find_element_by_xpath(
            './/td[4]').get_attribute('innerHTML'))
    return links, titulo, data


def elinput(Entrada, Xpath, web):
    element = web.find_element_by_xpath(Xpath)
    element.click()
    element.clear()
    element.send_keys(Entrada)
    web.find_element_by_xpath(
        '//*[@id="Form1"]/section[2]/div').click()  # Alterar foco


def porcentagem(valor, total):
    porcentagem = (valor/total)*100
    return int(porcentagem)


def start(P_chave, Data, hoje):
    links = []
    titulo = []
    data = []
    i = 0
    print("----[Script Iniciado]----")
    logging.warning(str(parse(datetime.now().isoformat(timespec='seconds'))) + ': ----[Script Iniciado]----')
    print("  Buscando de: " + datainicio + " até " + hoje)
    os.environ['WDM_LOG_LEVEL'] = '0'  # remove logs
    options = webdriver.ChromeOptions()  # remove logs
    options.add_experimental_option('excludeSwitches', ['enable-logging'])  # remove logs
    options.add_argument("--headless")  # remove logs
    os.environ['WDM_PRINT_FIRST_LINE'] = 'False'  # remove logs

    web = webdriver.Chrome(ChromeDriverManager(log_level=0).install(), options=options)
    web.get('http://diariooficial.rn.gov.br/dei/dorn3/Search.aspx')
    time.sleep(2)
    # Preenchendo palavra chave
    elinput(P_chave, '//*[@id="input-bs-keyword"]', web)
    elinput(Data, '//*[@id="input-bs-data"]', web)  # Preenchendo data inicio
    # Preenchimento da tada final (dia em que o script roda)
    elinput(hoje, '//*[@id="input-bs-data-2"]', web)
    print("----[Carregando o Portal]----", end="\r", flush=True)
    logging.warning(str(parse(datetime.now().isoformat(timespec='seconds'))) + ': ----[Carregando o Portal]----')
    Submit = web.find_element_by_xpath('//*[@id="submit-busca-simples"]')
    Submit.click()
    try:
      page = int(web.find_element_by_xpath('//*[@id="lblPagina"]').get_attribute('innerHTML')[12:])
    except: 
      print("Nenhuma matéria encontrada.          ")
      web.close()
      raise SystemExit(0)
    logging.warning(str(parse(datetime.now().isoformat(timespec='seconds'))) + ': ----[Capturando Links]----')
    while True:
        print("----[Capturando Links]---- " + str(porcentagem(i, page)) + "%" + " da captura de links" + " (Pagina " + str(i+1) + " de " + str(page) + ").", end="\r", flush=True)
        linkspage, titulopage, datapage = get_all_links(web)
        for j in range(0, len(linkspage)):
            links.append(linkspage[j])
            titulo.append(titulopage[j])
            data.append(datapage[j])
        if(i == page-1):
            pagesaida = web.find_element_by_xpath('//*[@id="lblPagina"]')
            break
        else:
            t = web.find_element_by_xpath(
                '//*[@id="Form1"]/section[2]/div/div[2]/a[2]')
            t.click()
            i += 1
    return links, titulo, data, web


def informacoes(links, titulo, data, web):
    linkslei = []
    titulolei = []
    datalei = []
    logging.warning(str(parse(datetime.now().isoformat(timespec='seconds'))) + ': ----[Procurando por Ocorrencias]----')
    for i in range(len(links)):
        print("----[Procurando por Ocorrencias]---- " + str(porcentagem(i,
              len(links))) + "%" + " das paginas                                         ", end="\r", flush=True)
        web.get(links[i])
        time.sleep(2)
        conteud = web.find_elements_by_class_name("WordSection1")
        for element in conteud:
            if ((element.text.find("14.133")) != -1) or ((element.text.find("14133")) != -1):
                linkslei.append(links[i])
                titulolei.append(titulo[i])
                datalei.append(data[i])
    web.close()
    return linkslei, titulolei, datalei


def repetido(elementos, comparar):
    index = 0
    for i in range(len(elementos)):
        if elementos[i] == comparar:
            index = i+1
    return index


def gerarExcel(links, titulo, data, name):
   print("----[Salvando documento]----", end="\r", flush=True)
   logging.warning(str(parse(datetime.now().isoformat(timespec='seconds'))) + ': ----[Salvando documento]----')
   hoje = date.today().strftime("%d/%m/%Y")
   cwd = os.getcwd()
   path = os.path.join(cwd, "resultado")
   if os.path.exists(path):
      if os.path.isfile(os.path.join(path, name)):
         wb = load_workbook(filename="resultado/" + name)
         ws = wb.active
         indexrepetido = repetido(links, ws.cell(1, 4).value)
         linksunicos = links[indexrepetido:]
         titulounicos = titulo[indexrepetido:]
         dataunicos = data[indexrepetido:]
         if linksunicos == []:
            ws.cell(1, 1, "Ultima verificação: " + hoje)
            ws.cell(1, 5, hoje)
            wb.save("resultado/" + name)
            print("Nao foram encontrados documentos novos nessa pesquisa                                ")
            logging.warning(str(parse(datetime.now().isoformat(timespec='seconds'))) + ': Nao foram encontrados documentos novos nessa pesquisa                                 ')
         else:
               increment = ws.cell(1, 2).value
               print("Foram encontrados " + str(len(linksunicos)) + " novo(s) documento(s) a partir de " + str(increment))
               logging.warning(str(parse(datetime.now().isoformat(timespec='seconds'))) + ': Foram encontrados ' + str(len(linksunicos)) + ' novo(s) documento(s)                        ')
               if increment and (ws.cell(1, 2, increment).value):
                  for row_num, dados in enumerate(linksunicos):
                     # Rownum percorre as lista, enquanto increment move o começo para o ultimo elemento da lista, e o +1 para pular o ultimo elemento
                     ws.cell(row_num+increment+1, 1, titulounicos[row_num])
                     ws.cell(row_num+increment+1, 2, dados)  # mesmo de cima
                     ws.cell(row_num+increment+1, 3, dataunicos[row_num])  # mesmo de cima
                     ws.cell(1, 2, row_num+increment+1)
                     ws.cell(1, 3, titulounicos[row_num])
                     ws.cell(1, 4, dados)
                     ws.cell(1, 5, dataunicos[row_num])
                  ws.cell(1, 1, "Ultima verificação: " + hoje)
                  wb.save("resultado/" + name)

               elif links == []:
                  ws.cell(1, 1, "Ultima verificação: " + hoje)
                  wb.save("resultado/" + name)
                  print("Nao foram encontrados documentos novos nessa pesquisa                  ")
                  logging.warning(str(parse(datetime.now().isoformat(timespec='seconds'))) + ': Nao foram encontrados documentos novos nessa pesquisa                      ')

               else:
                  ws.cell(1, 1, "Ultima verificação: " + hoje)
                  wb.save("resultado/" + name)
                  print("A planilha está vazia. Espere até o programa achar algum documento")
                  logging.warning(str(parse(datetime.now().isoformat(timespec='seconds'))) + ': A planilha está vazia. Espere até o programa achar algum documento                ')

      else:
         print("Foram encontrados " + str(len(links)) + " novo(s) documento(s)                              ")
         logging.warning(str(parse(datetime.now().isoformat(timespec='seconds'))) + ': Foram encontrados ' + str(len(links)) + ' novo(s) documento(s)                  ')
         with xlsxwriter.Workbook('resultado' + '/' + name) as workbook:
               worksheet = workbook.add_worksheet()
               worksheet.set_column('A:A', 40)
               worksheet.set_column('B:B', 40)
               worksheet.set_column('C:C', 20)
               worksheet.write_string(0, 0, "Ultima verificação: " + hoje)
               for row_num, dados in enumerate(links):
                  worksheet.write_string(row_num + 1, 0, titulo[row_num])
                  worksheet.write_string(row_num + 1, 1, dados)
                  worksheet.write_string(row_num + 1, 2, data[row_num])
                  # +1 por conta da mudança de coordenada (0:0) para (1:1) +1 para eliminar cabeçalho
                  worksheet.write(0, 1, row_num + 2)
                  worksheet.write_string(0, 2, titulo[row_num])
                  worksheet.write_string(0, 3, dados)
                  worksheet.write_string(0, 4, data[row_num])


def mkdir(linkslei, titulolei, datalei, name):
    cwd = os.getcwd()
    path = os.path.join(cwd, "resultado")
    if os.path.exists(path):
        if os.path.isdir(path):
            gerarExcel(linkslei, titulolei, datalei, name)
    else:
         os.mkdir("resultado")
         gerarExcel(linkslei, titulolei, datalei, name)


def datainicio(name):
    datapadrao = "01/04/2021"
    cwd = os.getcwd()
    path = os.path.join(cwd, "resultado")
    if os.path.exists(path):
        if os.path.isfile(os.path.join(path, name)):
            wb = load_workbook(filename="resultado/" + name)
            ws = wb.active
            if ws.cell(1, 5).value:
                return str(ws.cell(1, 5).value)
            else:
                return datapadrao
        else:
            return datapadrao
    else:
        return datapadrao


if __name__ == '__main__':
    links = []
    linkslei = []
    titulolei = []
    datalei = []
    titulo = []
    data = []
    name = 'Automação DO (Lei 14133).xlsx'
    datainicio = datainicio(name)
    hoje = date.today().strftime("%d/%m/%Y")
    cwd = os.getcwd()
    path = os.path.join(cwd, "logs")
    if os.path.exists(path):
        if os.path.isdir(path):
           logging.basicConfig(filename='logs/Log ' + date.today().strftime("%d-%m-%Y") + '.log', level=logging.WARNING)
        else :
           os.mkdir("logs")
           logging.basicConfig(filename='logs/Log ' + date.today().strftime("%d-%m-%Y") + '.log', level=logging.WARNING)
    else :
       os.mkdir("logs")
       logging.basicConfig(filename='logs/Log ' + date.today().strftime("%d-%m-%Y") + '.log', level=logging.WARNING)
    links, titulo, data, web = start("14.133", datainicio, hoje)  # parametros: palavra de pesquisa e numero de pag pesquisadas
    linkslei, titulolei, datalei = informacoes(links, titulo, data, web)
    mkdir(linkslei, titulolei, datalei, name)
    print("----[Concluido!]----")
    logging.warning(str(parse(datetime.now().isoformat(timespec='seconds'))) + ': ----[Concluido!]----')


# usar regex caso queira salvar o texto de um jeito diferente
